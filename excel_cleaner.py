import os
import re
from datetime import datetime, date
import openpyxl
import pandas as pd


def _enmascarar_credenciales(texto: str) -> str:
    """Enmascara contraseñas y claves de acceso en el texto procesado."""
    patron_pass = r'(?i)(contrase[ñn]a|password|pass|pwd|secret)\s*[:=]\s*([^\s\|]+)'
    return re.sub(patron_pass, r'\1: [PROTEGIDO]', texto)


def _formatear_valor_celda(val) -> str:
    """Formatea valores individuales de celdas eliminando NaNs, formateando enteros y fechas limpias."""
    if val is None or pd.isna(val):
        return "-"
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d %H:%M") if isinstance(val, datetime) and (val.hour or val.minute) else val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return f"{val:.2f}".rstrip('0').rstrip('.')
    
    val_str = str(val).strip()
    if val_str.lower() in ('none', 'nan', 'nat', 'null', '<na>', '', '#n/a', '#value!', '#ref!') or val_str.lower().startswith('unnamed:'):
        return "-"
    return val_str.replace("\n", " ").replace("|", "/")


def _dataframe_a_markdown(df: pd.DataFrame) -> str:
    """Convierte un DataFrame a Markdown limpio sin requerir tabulate, con formato de celdas pulido."""
    if df.empty:
        return ""
    
    cols = [_formatear_valor_celda(c) for c in df.columns]
    cols = [c if c != "-" and not c.lower().startswith("unnamed:") else f"Col_{i+1}" for i, c in enumerate(cols)]
    
    header = "| " + " | ".join(cols) + " |"
    separator = "| " + " | ".join([":---"] * len(cols)) + " |"
    rows = []
    
    for _, row in df.iterrows():
        valores_limpios = [_formatear_valor_celda(val) for val in row]
        # Omitir filas totalmente vacias o que solo tienen guiones
        if all(v == "-" for v in valores_limpios):
            continue
        row_str = "| " + " | ".join(valores_limpios) + " |"
        rows.append(_enmascarar_credenciales(row_str))
        
    if not rows:
        return ""
    return "\n".join([header, separator] + rows)


def _procesar_hoja_formulario(data: list) -> list:
    """Procesa hojas tipo formulario/ficha técnica (celdas combinadas, pares clave-valor dispersos)."""
    lineas = []
    for row in data:
        non_empty = []
        for c in row:
            fval = _formatear_valor_celda(c)
            if fval != "-" and not fval.lower().startswith("unnamed:"):
                non_empty.append(fval)
                
        if not non_empty:
            continue
        if len(non_empty) == 1:
            val = non_empty[0]
            if len(val) > 25 or any(kw in val.lower() for kw in ('balancer', 'servidor', 'datacenter', 'infraestructura', 'switch', 'vcloud', 'prtg', 'monitoreo')):
                lineas.append(f"\n### {val}\n")
            else:
                lineas.append(f"**{val}**")
        elif len(non_empty) == 2:
            lineas.append(_enmascarar_credenciales(f"* **{non_empty[0]}:** {non_empty[1]}"))
        else:
            # Agrupar de a pares (Clave -> Valor)
            pares = []
            i = 0
            while i < len(non_empty):
                if i + 1 < len(non_empty):
                    pares.append(f"**{non_empty[i]}:** {non_empty[i+1]}")
                    i += 2
                else:
                    pares.append(f"{non_empty[i]}")
                    i += 1
            lineas.append(_enmascarar_credenciales(f"* {' | '.join(pares)}"))
    return lineas


def procesar_excel_limpio(filepath: str) -> str:
    """
    Convierte archivos Excel (.xlsx, .xls) a Markdown limpio y estructurado.
    Detecta automáticamente si la hoja es una Matriz Tabular o una Ficha Técnica/Formulario.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return f"Error al abrir archivo Excel: {str(e)}"

    doc_md = []
    fname = os.path.basename(filepath)
    doc_md.append(f"# Inventario y Documentación Técnica: {fname}\n")

    for sname in wb.sheetnames:
        ws = wb[sname]
        data = []
        for row in ws.iter_rows(values_only=True):
            cleaned_row = [_formatear_valor_celda(cell) for cell in row]
            if any(c != "-" for c in cleaned_row):
                data.append([c if c != "-" else None for c in cleaned_row])

        if not data:
            continue

        df_raw = pd.DataFrame(data)
        df_raw = df_raw.dropna(axis=1, how='all')
        df_raw = df_raw.dropna(axis=0, how='all')

        if df_raw.empty or len(df_raw.columns) == 0:
            continue

        sheet_title = sname.strip()
        doc_md.append(f"\n## Hoja: {sheet_title}\n")

        # Calcular densidad de celdas no vacías
        total_cells = df_raw.size
        non_null_cells = df_raw.notnull().sum().sum()
        densidad = non_null_cells / total_cells if total_cells > 0 else 0

        # Si la densidad es baja (< 35%) o el nombre es MENU/FICHA, es un Formulario/Ficha técnica
        es_formulario = densidad < 0.35 or any(k in sheet_title.lower() for k in ('menu', 'ficha', 'mapa', 'general'))

        if es_formulario:
            lineas_form = _procesar_hoja_formulario(data)
            doc_md.extend(lineas_form)
            doc_md.append("")
        else:
            # Es una Matriz Tabular (inventario, servidores, usuarios, escalamientos)
            header_idx = 0
            max_headers = 0
            for idx in range(min(10, len(df_raw))):
                row_vals = df_raw.iloc[idx].dropna().tolist()
                str_count = sum(1 for v in row_vals if isinstance(v, str) and len(v) > 1 and not v.replace('.', '').isdigit())
                if str_count > max_headers:
                    max_headers = str_count
                    header_idx = idx

            raw_headers = df_raw.iloc[header_idx].tolist()
            df_table = df_raw.iloc[header_idx + 1:].copy()

            cols_validas = []
            clean_headers = []
            for i, h in enumerate(raw_headers):
                serie = df_table.iloc[:, i]
                if h is not None or serie.notnull().any():
                    cols_validas.append(i)
                    h_name = str(h).strip() if h is not None else f"Col_{len(clean_headers)+1}"
                    clean_headers.append(h_name)

            df_table = df_table.iloc[:, cols_validas]
            df_table.columns = clean_headers
            df_table = df_table.fillna('-')

            df_table = df_table[~(df_table == '-').all(axis=1)]

            if not df_table.empty:
                md_tabla = _dataframe_a_markdown(df_table)
                if md_tabla.strip():
                    doc_md.append(md_tabla)
                    doc_md.append("")

    return "\n".join(doc_md)
